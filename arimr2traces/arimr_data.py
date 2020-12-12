# -*- coding: utf-8 -*-
"""
Routines to get data from ARiMR website
"""

import hashlib
import requests
import codecs
from lxml import html
from datetime import datetime
from typing import List
from openpyxl import Workbook
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.utils import get_column_letter

from .animal import *


URL_WITH_PARAM = 'https://giw.doplaty.gov.pl/%s'
MAIN_URL = 'https://giw.doplaty.gov.pl/index.php'

headers = {
    'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64; rv:58.0) Gecko/20100101 Firefox/58.0',
    'Referer': 'https://giw.doplaty.gov.pl/',
    'Accept-Language': 'en-US,en;q=0.5'
}


def get_texts_list_from_tags(tags):
    return [''.join(t.itertext()).strip() for t in tags]


class ArimrDataExtractor(object):
    """
    Class to extract data from ARiMR website and based on this creates Animal objects

    """

    def __init__(self):
        self.login_payload = {}
        self.animal_list = list()
        self.session = requests.Session()
        self.logged_in = False

        self.create_animal_events_calls = {
            ANIMAL_SPECIES_CATTLE: self._create_animal_events_for_cattle,
            ANIMAL_SPECIES_SHEEP: self._create_animal_events_for_sheep
        }


    def login(self, username, password):
        """
        login to website

        :param username:
        :param password:
        :return:
        """
        self.login_payload = {
            'USR_NAME': hashlib.md5(username.encode('utf-8')).hexdigest(),
            'USR_PSW': hashlib.md5(password.encode('utf-8')).hexdigest()
        }

        r = self.session.post(MAIN_URL, data=self.login_payload, headers=headers)
        self.logged_in = True if 'Wyloguj' in r.text else False

        return self.logged_in

    def _get_passport_latest_data(self, link):
        """
        extracts passport data

        :param link:
        :return:
        """
        passport_id, passport_status = '', ''
        r, tree = self._webpage_get(link)
        if r.status_code == 200:
            table_tags = tree.xpath("//table[@id='opcja']/tr//td")
            passport_data_texts = get_texts_list_from_tags(table_tags)
            tdp_count = len(passport_data_texts)
            if tdp_count > 16:
                passport_id = passport_data_texts[12]
                passport_status = passport_data_texts[17]
        return passport_id, passport_status

    def _create_animal_events_for_cattle(self, animal_data_texts):
        """
        gets cattle's events

        :param animal_data_texts:
        :return:
        """
        i = 9
        while i < len(animal_data_texts):
            animal_event = AnimalEvent(typ_zdarzenie=animal_data_texts[i + 1], nr_w_oznakowaniu=animal_data_texts[i + 2],
                                       nr_siedziby=animal_data_texts[i + 3], data_zdarzenia=animal_data_texts[i + 4],
                                       stan_zdarzenia=animal_data_texts[i + 5],
                                       # liczba_zwierzat_w_zgloszeniu=animal_data_texts[i + 6],
                                       nr_siedziby_stad_kompl=animal_data_texts[i + 7], uwagi=animal_data_texts[i + 8])
            yield animal_event
            i += 9

    def _create_animal_events_for_sheep(self, animal_data_texts):
        """
        gets sheep's events

        :param animal_data_texts:
        :return:
        """
        i = 8
        while i < len(animal_data_texts):
            animal_event = AnimalEvent(typ_zdarzenie=animal_data_texts[i + 1], nr_w_oznakowaniu=animal_data_texts[i + 2],
                nr_siedziby=animal_data_texts[i + 3], data_zdarzenia=animal_data_texts[i + 4],
                stan_zdarzenia=animal_data_texts[i + 5], nr_siedziby_stad_kompl=animal_data_texts[i + 6],
                uwagi=animal_data_texts[i + 7])
            yield animal_event
            i += 8

    def _webpage_get(self, url, **kwargs):
        r = self.session.get(url, **kwargs)
        tree = html.fromstring(r.content)
        return r, tree

    def _get_animal_data(self, animal_id):
        """
        main function to get(extarct) animal data from ARiMR webpage

        :param animal_id:
        :return:
        """
        params = {'zmNumer': animal_id, 'OPX': '1'}
        r, tree = self._webpage_get(MAIN_URL, params=params)
        animal = Animal(animal_id)
        if not r.status_code == 200:
            animal.process_status += ' Kod błędu: %d' % r.status_code
            return animal
        # get link point to Animal Events (Zdarzenia)
        links_list = tree.xpath('/html/body/table/tr[3]/td/table/tr[1]/td/table/table/tr[1]/td/table[2]/tr[2]/td//a/@href')
        # Tabela zdarzen OPX=1&LST=2
        events_table_tag = [link for link in links_list if 'OPX=1&LST=2' in link]
        # Tabela paszportu OPX=1&LST=1
        passport_table_tag = [link for link in links_list if 'OPX=1&LST=1' in link]
        if len(events_table_tag) > 0:
            # zdarzenia pobrane
            r, tree = self._webpage_get(URL_WITH_PARAM % events_table_tag[0])
            animal_data_tags = tree.xpath("/html/body/table/tr[3]/td/table/tr[1]/td/table/table/tr[1]/td/div[@id='raport']/table[2]//tr/td[3]")
            animal_data_list = get_texts_list_from_tags(animal_data_tags)
            # zdarzenia zwierzaka
            animal_events_tags = tree.xpath("//div[@id='opcja']/table[2]/tr//td")
            events_data_texts = get_texts_list_from_tags(animal_events_tags)
            # print(events_data_texts)
            herd_id = ''
            herd_id_tag = tree.xpath("/html/body/table/tr[3]/td/table/tr[1]/td/table/tr[5]/td/table/tr[1]/td/a[2]/text()")
            if herd_id_tag:
                herd_id = herd_id_tag[0].replace('Siedziba stada ', '')
            #
            passport_id, passport_status = self._get_passport_latest_data(URL_WITH_PARAM % passport_table_tag[0])
            #
            try:
                animal.set_data(animal_data_list[1], animal_data_list[2], animal_data_list[3],
                                  animal_data_list[4], animal_data_list[5], animal_data_list[6], animal_data_list[7],
                                  animal_data_list[8], animal_data_list[9], animal_data_list[10], animal_data_list[11],
                                  animal_data_list[12], animal_data_list[13], passport_id, passport_status, herd_id)
                #
                for animal_event in self.create_animal_events_calls[animal.species](events_data_texts):
                    animal.events_list.append(animal_event)
                animal.process_status = ANIMAL_OK
            except:
                pass
        return animal

    def get_animals_from_ids_list(self, animals_ids_list):
        """
        Main function to retrieve animal data

        :param animals_ids_list:
        :return:
        """
        for animal_id in animals_ids_list:
            yield self._get_animal_data(animal_id)


def _save_report_file_as_csv(report_file_name, dt_of_fetching_data, header, animals_list, now):
    fd = codecs.open(report_file_name, 'w', 'utf-8-sig')
    try:
        writer = csv.writer(fd, dialect='excel', quotechar='"', quoting=csv.QUOTE_ALL)
        writer.writerow(['Data pobrania danych: %s' % dt_of_fetching_data])
        writer.writerow(header)
        i = 1
        for animal in animals_list:
            row = [i, animal.herd_id, animal.animal_id, animal.process_status, animal.passport_id,
                   animal.passport_status,
                   animal.age_in_months(now), animal.species, animal.sex, animal.breed,
                   animal.typ_uzyt, animal.status, animal.birth_date, animal.get_animal_events_states_text()]
            #
            writer.writerow(row)
            i += 1
    finally:
        fd.close()


def _save_report_file_as_excel(report_file_name, dt_of_fetching_data, header, animals_list, now):
    wb = Workbook()
    ws = wb.active
    row = 1
    col = 1
    # date
    ws.cell(column=1, row=row, value='Data pobrania danych: %s' % dt_of_fetching_data)
    # header
    row += 1
    for header_elem in header:
        ws.cell(column=col, row=row, value=header_elem)
        col += 1
    # values
    row += 1
    cnt = 1
    for animal in animals_list:
        col = 1
        values = [cnt, animal.herd_id, animal.animal_id, animal.process_status, animal.passport_id,
               animal.passport_status,
               animal.age_in_months(now), animal.species, animal.sex, animal.breed,
               animal.typ_uzyt, animal.status, animal.birth_date, animal.get_animal_events_states_text()]
        for value in values:
            ws.cell(column=col, row=row, value=value)
            col += 1
        row += 1
        cnt += 1

    for col in range(1, 10):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter] = ColumnDimension(ws, customWidth=True)
        ws.column_dimensions[col_letter].width = 24

    wb.save(report_file_name)


def save_report_file(report_file_name, animals_list: List[Animal], output_format='excel'):
    """
    Saves report file in CSV with needed data.

    :param report_file_name:
    :param animals_list:
    :param output_format:

    :return:
    """
    now = datetime.now()
    dt_of_fetching_data = now.strftime('%Y-%m-%d %H:%M:%S')
    header = ['Lp', 'Nr stada', 'Nr identyfikatora', 'Status pobierania danych', 'Nr paszportu', 'Status paszportu',
                       'Wiek (m)', 'Gatunek',
                       'Plec', 'Rasa', 'Typ użytkowy', 'Status', 'Data urodzenia', 'Stan zdarzeń']
    if output_format == 'excel':
        _save_report_file_as_excel(report_file_name, dt_of_fetching_data, header, animals_list, now)
    else:
        _save_report_file_as_csv(report_file_name, dt_of_fetching_data, header, animals_list, now)


def save_traces_file(traces_file_name, animals_list: List[Animal]):
    """
    Saves TRACES compatible import file, based on animals in animals_list

    :param traces_file_name:
    :param animals_list:
    :return:
    """
    def save_traces_file_cattles():
        fd = open(traces_file_name, 'w')
        try:
            writer = csv.writer(fd, dialect='excel', delimiter=';', quoting=csv.QUOTE_NONE)
            writer.writerow(['[COLUMNS]'])
            writer.writerow(['official_ident', 'numpassportemp'])
            writer.writerow([])
            writer.writerow(['[DATA]'])
            for animal in animals_list:
                row = [animal.animal_id, animal.passport_id]
                writer.writerow(row)
            writer.writerow(['[DATA]'])
        finally:
            fd.close()

    def save_traces_file_sheeps():
        fd = open(traces_file_name, 'w')
        try:
            writer = csv.writer(fd, dialect='excel', delimiter=';', quoting=csv.QUOTE_NONE)
            writer.writerow(['[COLUMNS]'])
            writer.writerow(['official_indiv', 'age_month', 'sexinfo'])
            writer.writerow([''])
            writer.writerow(['[DATA]'])
            now = datetime.now()
            for animal in animals_list:
                row = [animal.animal_id, animal.age_in_months(now), animal.get_traces_sex()]
                writer.writerow(row)
            writer.writerow(['[DATA]'])
        finally:
            fd.close()

    if not animals_list:
        return False

    if animals_list[0].species == ANIMAL_SPECIES_CATTLE:
        save_traces_file_cattles()
    elif animals_list[0].species == ANIMAL_SPECIES_SHEEP:
        save_traces_file_sheeps()
    else:
        raise NotImplementedError

    return True
